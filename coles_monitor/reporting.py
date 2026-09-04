from email.message import EmailMessage
from email.utils import formatdate
from html import escape
import smtplib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from .matcher import category_group


HEADERS = ["Observed (UTC)", "Retailer", "Brand", "Product", "Size", "Change Summary",
           "Current Price (AUD)", "Original Price (AUD)", "Discount", "Availability",
           "Product URL", "Event ID"]
RETAILERS = ("Coles", "Woolworths")
GROUPS = ("Tomato Paste", "Pasta Sauce", "Passata", "Pesto")
PRODUCT_HEADERS = ["Product ID", "Brand", "Product", "Size", "Change Summary",
                   "Current Price (AUD)", "Original Price (AUD)", "Discount",
                   "Availability", "Product URL"]


def _group_for(item):
    return item.get("category_group") or category_group(item.get("name", ""))


def _sort_key(item):
    return (str(item.get("brand", "")).casefold(), str(item.get("name", "")).casefold())


def _workbook_current_price(item):
    price = item.get("price")
    promotion = item.get("promotional_price")
    if isinstance(promotion, str):
        value = f"{_format_price(price)} each — {promotion}"
    else:
        value = price
    if item.get("online_only") and promotion is not None:
        return f"{_format_price(value)} (Online only promotion)" if not isinstance(value, str) \
            else value + " (Online only promotion)"
    return value


def _style_sheet(ws, widths):
    header_fill = PatternFill("solid", fgColor="C41230")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False


def _write_retailer_sheet(wb, retailer, report_events, failed=False):
    ws = wb.create_sheet(retailer, len(wb.sheetnames) - 1)
    last_column = get_column_letter(len(PRODUCT_HEADERS))
    ws.merge_cells(f"A1:{last_column}1")
    ws["A1"] = retailer
    ws["A1"].fill = PatternFill("solid", fgColor="C41230")
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")
    if failed:
        ws.merge_cells(start_row=3, start_column=1, end_row=3,
                       end_column=len(PRODUCT_HEADERS))
        ws.cell(3, 1, "Refresh unavailable; the last verified snapshot was retained. "
                      "No no-change conclusion was made for this retailer.")
        ws.cell(3, 1).fill = PatternFill("solid", fgColor="FFF2CC")
        ws.cell(3, 1).font = Font(bold=True, color="9C6500")
        ws.cell(3, 1).alignment = Alignment(wrap_text=True)
        for index, width in enumerate([18, 18, 50, 14, 22, 28, 18, 12, 22, 52], 1):
            ws.column_dimensions[get_column_letter(index)].width = width
        ws.sheet_view.showGridLines = False
        return
    row = 3
    retailer_products = [event for event in report_events if event.get("retailer") == retailer]
    for group_index, group in enumerate(GROUPS, 1):
        ws.merge_cells(start_row=row, start_column=1, end_row=row,
                       end_column=len(PRODUCT_HEADERS))
        ws.cell(row, 1, group)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor="E7E6E6")
        ws.cell(row, 1).font = Font(bold=True, size=12)
        row += 1
        header_row = row
        ws.append(PRODUCT_HEADERS)
        for cell in ws[header_row]:
            cell.fill = PatternFill("solid", fgColor="7F6000")
            cell.font = Font(color="FFFFFF", bold=True)
        matches = sorted((event for event in retailer_products
                          if _group_for(event) == group), key=_sort_key)
        for product in matches:
            ws.append([product.get("product_id", ""), product.get("brand", ""), product["name"],
                       product["size"], product.get("change_type", ""),
                       _workbook_current_price(product), product.get("original_price"),
                       product.get("discount_percent"), product.get("availability_label", ""),
                       product["product_url"]])
            data_row = ws.max_row
            ws.cell(data_row, 3).hyperlink = product["product_url"]
            ws.cell(data_row, 3).style = "Hyperlink"
            for column in (6, 7):
                ws.cell(data_row, column).number_format = '"$"0.00'
            ws.cell(data_row, 8).number_format = "0.0%"
        if matches:
            table_name = f'{retailer}{group.replace(" ", "")}Table'
            table = Table(displayName=table_name,
                          ref=f"A{header_row}:{last_column}{ws.max_row}")
            table.tableStyleInfo = TableStyleInfo(name=f"TableStyleMedium{group_index + 1}",
                                                  showRowStripes=True)
            ws.add_table(table)
        else:
            row += 1
            ws.cell(row, 1, "No matching SKUs")
            ws.cell(row, 1).font = Font(italic=True, color="666666")
        row = ws.max_row + 2
    for index, width in enumerate([18, 18, 50, 14, 22, 28, 18, 12, 22, 52], 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False


def write_workbook(path, events, current=None, report_events=None, failures=()):
    wb = Workbook()
    ws = wb.active
    ws.title = "Change History"
    ws.append(HEADERS)
    for event in events:
        ws.append([event["observed_at"], event.get("retailer", ""), event.get("brand", ""),
                   event["name"], event["size"], event["change_type"],
                   _workbook_current_price(event), event.get("original_price"),
                   event.get("discount_percent"), event.get("availability_label", ""),
                   event["product_url"], event["event_id"]])
        row = ws.max_row
        ws.cell(row, 4).hyperlink = event["product_url"]
        ws.cell(row, 4).style = "Hyperlink"
    _style_sheet(ws, [22, 14, 18, 48, 14, 22, 28, 18, 12, 22, 52, 28])
    for column in ("G", "H"):
        for cell in ws[column][1:]:
            cell.number_format = '"$"0.00'
    for cell in ws["I"][1:]:
        cell.number_format = "0.0%"
    if current is not None:
        failed_retailers = {str(failure).split(":", 1)[0] for failure in failures}
        for retailer in RETAILERS:
            _write_retailer_sheet(wb, retailer, report_events or [],
                                  retailer in failed_retailers)
    wb.save(path)


def _sectioned_html(items, row_renderer, failures=()):
    sections = []
    failed_retailers = {str(failure).split(":", 1)[0] for failure in failures}
    for retailer in RETAILERS:
        sections.append(f"<h2>{retailer}</h2>")
        if retailer in failed_retailers:
            sections.append("<p><strong>Refresh unavailable.</strong> The last verified "
                            "snapshot was retained, so no no-change conclusion was made "
                            "for this retailer.</p>")
            continue
        retailer_items = [item for item in items if item.get("retailer") == retailer]
        for group in GROUPS:
            grouped = sorted((item for item in retailer_items
                              if _group_for(item) == group), key=_sort_key)
            sections.append(f"<h3>{group}</h3>")
            sections.append(row_renderer(grouped))
    return "".join(sections)


def _format_price(value):
    if value is None:
        return ""
    try:
        return f'${float(value):.2f}'
    except (TypeError, ValueError):
        return escape(str(value))


def _format_current_price(item):
    price = _format_price(item.get("price"))
    promotion = item.get("promotional_price")
    if isinstance(promotion, str):
        price = f"{price} each — {escape(promotion)}"
    if item.get("online_only") and promotion is not None:
        price += " (Online only promotion)"
    return price


def email_visible_events(events):
    return [event for event in events if not event.get("promotion_ended")]


def render_html(events, failures=()):
    events = email_visible_events(events)
    def render_table(grouped):
        if not grouped:
            return "<p>No changes.</p>"
        rows = []
        for e in grouped:
            name = f'<a href="{escape(e["product_url"], quote=True)}">{escape(e["name"])}</a>'
            price = _format_current_price(e)
            original = _format_price(e.get("original_price"))
            discount = "" if e.get("discount_percent") is None else f'{float(e["discount_percent"]):.1%}'
            rows.append("<tr>" + "".join(f"<td>{v}</td>" for v in
                        [escape(e.get("brand", "")), name, escape(e["size"]),
                         escape(e["change_type"]), price, original, discount,
                         escape(e.get("availability_label", ""))]) + "</tr>")
        return """<table style="border-collapse:collapse" border="1" cellpadding="6"><thead><tr>
<th>Brand</th><th>Product</th><th>Size</th><th>Change Summary</th><th>Current Price</th>
<th>Original Price</th><th>Discount</th><th>Availability</th>
</tr></thead><tbody>""" + "".join(rows) + "</tbody></table>"
    return ("<!doctype html><html><body><p>Changes detected since the previous successful weekly run:</p>" +
            _sectioned_html(events, render_table, failures) +
            "<p>Sources: Coles and Woolworths product pages linked above.</p></body></html>")


def render_baseline_html(current, test=False):
    def render_table(grouped):
        if not grouped:
            return "<p>No matching SKUs.</p>"
        rows = []
        for product in grouped:
            name = f'<a href="{escape(product["product_url"], quote=True)}">{escape(product["name"])}</a>'
            price = _format_current_price(product)
            original = _format_price(product.get("original_price"))
            discount = ("" if product.get("discount_percent") is None else
                        f'{float(product["discount_percent"]):.1%}')
            rows.append("<tr>" + "".join(f"<td>{v}</td>" for v in
                        [escape(product.get("brand", "")), name, escape(product["size"]),
                         price, original, discount,
                         escape(product.get("availability_label", ""))]) + "</tr>")
        return """<table style="border-collapse:collapse" border="1" cellpadding="6"><thead><tr>
<th>Brand</th><th>Product</th><th>Size</th><th>Current Price</th><th>Original Price</th>
<th>Discount</th><th>Availability</th></tr></thead><tbody>""" + "".join(rows) + "</tbody></table>"
    intro = ("Live test baseline for Coles and Woolworths products requested for "
             "Cheltenham VIC 3192:" if test else
             "Initial Coles and Woolworths product baseline for Cheltenham VIC 3192:")
    return f"<!doctype html><html><body><p>{intro}</p>\n" + \
        _sectioned_html(list(current.values()), render_table) + \
        "<p>Future emails will contain only new changes.</p></body></html>"


def send_email(sender, recipient, app_password, events, workbook_path, baseline=None,
               failures=(), test=False):
    msg = EmailMessage()
    msg["From"], msg["To"] = sender, recipient
    msg["Date"] = formatdate(localtime=False)
    if baseline is not None:
        prefix = "TEST - " if test else ""
        msg["Subject"] = (f"{prefix}Coles & Woolworths product baseline - "
                          f"{len(baseline)} products")
        msg.set_content(("Live test baseline" if test else "Initial baseline") +
                        " for Coles and Woolworths products. Open as HTML or see "
                        "the attached Excel workbook.")
        msg.add_alternative(render_baseline_html(baseline, test=test), subtype="html")
    else:
        events = email_visible_events(events)
        if failures and not events:
            msg["Subject"] = "Coles & Woolworths monitor warning — refresh incomplete"
            msg.set_content("A retailer could not be refreshed. The last verified snapshot was retained.")
        else:
            msg["Subject"] = f"Coles & Woolworths product changes — {len(events)} change{'s' if len(events) != 1 else ''}"
            msg.set_content("Changes were detected. Open this message as HTML or see the attached Excel history.")
        msg.add_alternative(render_html(events, failures), subtype="html")
    with open(workbook_path, "rb") as handle:
        msg.add_attachment(handle.read(), maintype="application",
                           subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           filename="coles-woolworths-sauce-change-history.xlsx")
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=30) as smtp:
        smtp.login(sender, app_password)
        smtp.send_message(msg)

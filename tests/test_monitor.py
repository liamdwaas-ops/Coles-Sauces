import unittest

from coles_monitor.changes import compare, consolidate_events, visible_products
from coles_monitor.matcher import (category_group, is_allowed_product, is_wanted_name,
                                   keyword_group, split_name_size)
from coles_monitor.reporting import (email_visible_events, render_baseline_html,
                                     render_html, write_workbook)
from openpyxl import load_workbook
from pathlib import Path
from tempfile import TemporaryDirectory
from run_monitor import scrape_with_fallback
from coles_monitor.scraper import ColesScraper
from coles_monitor.woolworths import WoolworthsScraper


class MatcherTests(unittest.TestCase):
    def test_exact_rules(self):
        self.assertTrue(is_wanted_name("Brand Pasta Bake Sauce"))
        self.assertTrue(is_wanted_name("Brand Tomato Paste"))
        self.assertTrue(is_wanted_name("Brand Passata"))
        self.assertTrue(is_wanted_name("Brand Pesto Genovese"))
        self.assertFalse(is_wanted_name("Tomato Sauce"))
        self.assertFalse(is_wanted_name("Pasta Penne"))

    def test_title_and_brand_exclusions(self):
        self.assertFalse(is_allowed_product("Fresh Tomato Pasta Sauce", "Example"))
        self.assertFalse(is_allowed_product("Tomato Pasta Sauce", "Continental"))
        self.assertFalse(is_allowed_product("Tomato Paste", "Sirena"))
        self.assertTrue(is_allowed_product("Tomato Paste", "Leggo's"))
        self.assertFalse(is_allowed_product("Basil Pesto", "Rana"))
        self.assertFalse(is_allowed_product("Manual Food Chopper Pesto", "Example"))
        self.assertFalse(is_allowed_product("Pesto Throw Rug", "Example"))
        self.assertFalse(is_allowed_product("San Remo Tomato Paste", "Example"))
        self.assertFalse(is_allowed_product("My Muscle Chef Pasta Sauce", "Example"))

    def test_name_size(self):
        self.assertEqual(split_name_size("Brand Pesto | 190g"), ("Brand Pesto", "190g"))

    def test_exclusive_keyword_group_priority(self):
        self.assertEqual(keyword_group("Tomato Paste Passata"), "Tomato Paste")
        self.assertEqual(keyword_group("Passata Pasta Sauce"), "Pasta Sauce")
        self.assertEqual(keyword_group("Basil Pesto"), "Pesto")
        self.assertIsNone(keyword_group("Tomato Sauce"))

    def test_retailer_taxonomy_classifies_stir_through_as_pasta_sauce(self):
        name = "Leggo's Stir Through Sauce Roasted Vegetables"
        taxonomy = 'PASTA SAUCE & CHEESE ["Italian", "Pizza & Pasta Sauce"]'
        self.assertEqual(category_group(name, taxonomy), "Pasta Sauce")
        self.assertTrue(is_allowed_product(name, "Leggo's", taxonomy))


class ReportingTests(unittest.TestCase):
    def test_retailer_and_keyword_sections_do_not_duplicate_skus(self):
        current = {
            "coles:1": {"retailer": "Coles", "brand": "A", "name": "Tomato Paste Passata",
                        "price": 2.0, "size": "100g", "image_url": "", "product_url": "https://example/1"},
            "woolworths:2": {"retailer": "Woolworths", "brand": "B", "name": "Pasta Sauce",
                             "price": 3.0, "size": "500g", "image_url": "", "product_url": "https://example/2"},
            "coles:unchanged": {"retailer": "Coles", "brand": "C", "name": "Basil Pesto",
                                "price": 4.0, "size": "190g", "image_url": "",
                                "product_url": "https://example/3"},
        }
        changed = {key: value for key, value in current.items() if key != "coles:unchanged"}
        report_events = compare({}, changed, "2026-01-01T00:00:00+00:00")
        with TemporaryDirectory() as directory:
            path = Path(directory) / "report.xlsx"
            write_workbook(path, report_events, current, report_events=report_events)
            workbook = load_workbook(path)
            self.assertEqual(workbook.sheetnames, ["Coles", "Woolworths", "Change History"])
            self.assertNotIn("Image URL", [cell.value for cell in workbook["Coles"][4]])
            report_headers = [cell.value for cell in workbook["Coles"][4]]
            self.assertEqual(report_headers[2:5], ["Product", "Size", "Change Summary"])
            self.assertNotIn("Promotional Price (AUD)", report_headers)
            self.assertNotIn("Online Only", report_headers)
            history_headers = [cell.value for cell in workbook["Change History"][1]]
            self.assertNotIn("Before", history_headers)
            self.assertNotIn("After", history_headers)
            self.assertNotIn("Image URL", history_headers)
            self.assertNotIn("Promotional Price (AUD)", history_headers)
            self.assertNotIn("Online Only", history_headers)
            ids = []
            for sheet_name in ("Coles", "Woolworths"):
                ids.extend(cell.value for cell in workbook[sheet_name]["A"]
                           if isinstance(cell.value, str) and ":" in cell.value)
            self.assertCountEqual(ids, changed.keys())
            self.assertNotIn("coles:unchanged", ids)
        html = render_baseline_html(current)
        self.assertIn("<h2>Coles</h2>", html)
        self.assertIn("<h2>Woolworths</h2>", html)
        self.assertEqual(html.count(">Tomato Paste Passata</a>"), 1)
        self.assertNotIn("<th>Promotional Price</th>", html)
        self.assertNotIn("<th>Online Only</th>", html)
        self.assertIn("<th>Product</th><th>Size</th>", html)

    def test_email_orders_each_category_by_brand_and_marks_online_promotion(self):
        products = {
            "woolworths:1": {"retailer": "Woolworths", "brand": "Zulu",
                              "name": "Zulu Pasta Sauce", "size": "500g", "price": 4.0,
                              "product_url": "https://example/1", "availability_label": "Available"},
            "woolworths:2": {"retailer": "Woolworths", "brand": "Alpha",
                              "name": "Alpha Pasta Sauce", "size": "500g", "price": 3.0,
                              "original_price": 4.0, "promotional_price": 3.0,
                              "discount_percent": 0.25, "online_only": True,
                              "product_url": "https://example/2", "availability_label": "Available"},
        }
        html = render_baseline_html(products)
        self.assertLess(html.index("Alpha Pasta Sauce"), html.index("Zulu Pasta Sauce"))
        self.assertIn("$3.00 (Online only promotion)", html)

    def test_failed_retailer_is_not_described_as_no_changes(self):
        html = render_html([], failures=["Coles: ScrapeError: blocked"])
        coles_section = html.split("<h2>Coles</h2>", 1)[1].split("<h2>Woolworths</h2>", 1)[0]
        self.assertIn("Refresh unavailable", coles_section)
        self.assertNotIn("No changes", coles_section)


class ScrapeFallbackTests(unittest.TestCase):
    def test_failed_retailer_retains_verified_snapshot(self):
        class FailedScraper:
            def scrape(self, queries):
                raise RuntimeError("blocked")

        class WorkingScraper:
            def scrape(self, queries):
                return {"woolworths:2": {"retailer": "Woolworths", "name": "New Passata"}}

        previous = {
            "coles:1": {"retailer": "Coles", "name": "Verified Tomato Paste"},
            "woolworths:1": {"retailer": "Woolworths", "name": "Old Passata"},
        }
        current, failures = scrape_with_fallback(
            (("Coles", FailedScraper()), ("Woolworths", WorkingScraper())), [], previous
        )
        self.assertIn("coles:1", current)
        self.assertNotIn("woolworths:1", current)
        self.assertIn("woolworths:2", current)
        self.assertEqual(len(failures), 1)


class LocationTests(unittest.TestCase):
    def test_cheltenham_location_is_retained(self):
        location = {"suburb": "Cheltenham", "postcode": "3192", "state": "VIC",
                    "context_mode": "delivery"}
        scraper = ColesScraper(location=location)
        self.assertEqual(scraper.location, location)

    def test_coles_multibuy_text_is_captured(self):
        _, product = ColesScraper._product({
            "id": "1", "name": "Example Passata", "availability": True,
            "pricing": {"now": 4.6, "specialType": "MULTI_SAVE",
                        "offerDescription": "Pick any 2 for $7",
                        "multiBuyPromotion": {"minQuantity": 2, "reward": 3.5}},
        })
        self.assertEqual(product["original_price"], 4.6)
        self.assertEqual(product["promotional_price"], "Pick any 2 for $7")
        self.assertEqual(product["discount_percent"], 0.2391)

    def test_coles_multibuy_outside_pricing_is_captured(self):
        _, product = ColesScraper._product({
            "id": "2", "name": "Example Pasta Sauce", "availability": True,
            "pricing": {"now": 4.5, "specialType": "MULTI_SAVE"},
            "promotions": [{"offerDescription": "Any 2 for $7"}],
        })
        self.assertEqual(product["original_price"], 4.5)
        self.assertEqual(product["promotional_price"], "2 for $7")
        self.assertEqual(product["discount_percent"], 0.2222)

    def test_coles_multibuy_badge_is_captured(self):
        _, product = ColesScraper._product({
            "id": "3", "name": "Example Pesto", "availability": True,
            "pricing": {"now": 6.0},
            "badges": {"promotion": {"PromotionText": "Buy 2 for $10.00"}},
        })
        self.assertEqual(product["promotional_price"], "2 for $10.00")
        self.assertEqual(product["discount_percent"], 0.1667)

    def test_coles_ordered_images_are_retained(self):
        _, product = ColesScraper._product({
            "id": "4", "name": "Example Passata", "availability": True,
            "pricing": {"now": 3.0},
            "imageUris": [{"uri": "/4/4.jpg"}, {"uri": "/4/4_2.jpg"}],
        })
        self.assertEqual(product["image_urls"], [
            "https://cdn.productimages.coles.com.au/productimages/4/4.jpg",
            "https://cdn.productimages.coles.com.au/productimages/4/4_2.jpg",
        ])

    def test_coles_taxonomy_classifies_stir_through_as_pasta_sauce(self):
        _, product = ColesScraper._product({
            "id": "5", "name": "Roasted Vegetables Stir Through Sauce",
            "brand": "Leggo's", "availability": True, "pricing": {"now": 4.6},
            "merchandiseHeir": {
                "category": "MEAL BASES", "subCategory": "PASTA SAUCE",
                "className": "CHUNKY",
            },
            "onlineHeirs": [{"aisle": "Pizza & Pasta"}],
        })
        self.assertEqual(product["category_group"], "Pasta Sauce")


class WoolworthsTests(unittest.TestCase):
    def test_nested_search_response_mapping(self):
        payload = {"Products": [{"Products": [{
            "Stockcode": 502381,
            "Name": "Woolworths Passata 680g",
            "PackageSize": "680g",
            "Price": 2.25,
            "MediumImageFile": "https://cdn.example.test/502381.jpg",
            "UrlFriendlyName": "woolworths-passata"
        }]}]}
        products = WoolworthsScraper._find_products(payload)
        self.assertEqual(len(products), 1)
        product_id, product = WoolworthsScraper._product(products[0])
        self.assertEqual(product_id, "woolworths:502381")
        self.assertEqual(product["retailer"], "Woolworths")
        self.assertEqual(product["name"], "Woolworths Passata")
        self.assertEqual(product["size"], "680g")
        self.assertEqual(product["price"], 2.25)
        self.assertFalse(product["online_only"])

    def test_woolworths_taxonomy_and_ordered_images_are_retained(self):
        _, product = WoolworthsScraper._product({
            "Stockcode": 957033,
            "Name": "Leggo's Stir Through Tomato Garlic & Caramelised Onion Sauce",
            "Brand": "Leggo's", "Price": 4.3, "PackageSize": "350g",
            "MediumImageFile": "https://cdn.example/medium/957033.jpg",
            "AdditionalAttributes": {
                "sapsubcategoryname": "PASTA SAUCE & CHEESE",
                "sapsegmentname": "PASTA SAUCE STIR THRU",
                "productimages": "957033.jpg,957033_2.jpg",
            },
        })
        self.assertEqual(product["category_group"], "Pasta Sauce")
        self.assertEqual(product["image_urls"], [
            "https://cdn.example/medium/957033.jpg",
            "https://cdn.example/medium/957033_2.jpg",
        ])

    def test_woolworths_online_only_flag(self):
        _, product = WoolworthsScraper._product({
            "Stockcode": 99, "Name": "Example Pesto 190g", "PackageSize": "190g",
            "Price": 4.0, "IsOnlineOnly": True
        })
        self.assertTrue(product["online_only"])

    def test_woolworths_promotion_fields_require_explicit_promo(self):
        _, promo = WoolworthsScraper._product({
            "Stockcode": 1, "Name": "Example Passata 700g", "PackageSize": "700g",
            "Brand": "Example", "Price": 3.0, "WasPrice": 4.0, "IsOnSpecial": True,
            "IsAvailable": True, "IsInStock": True
        })
        self.assertEqual(promo["original_price"], 4.0)
        self.assertEqual(promo["promotional_price"], 3.0)
        self.assertEqual(promo["discount_percent"], 0.25)
        _, not_promo = WoolworthsScraper._product({
            "Stockcode": 2, "Name": "Example Passata 700g", "PackageSize": "700g",
            "Brand": "Example", "Price": 3.0, "WasPrice": 4.0, "IsOnSpecial": False
        })
        self.assertIsNone(not_promo["original_price"])

    def test_woolworths_multibuy_text_is_captured(self):
        _, product = WoolworthsScraper._product({
            "Stockcode": 5, "Name": "Example Passata", "Price": 4.0,
            "PromotionDescription": "2 for $6", "IsAvailable": True, "IsInStock": True,
        })
        self.assertEqual(product["original_price"], 4.0)
        self.assertEqual(product["promotional_price"], "2 for $6")
        self.assertEqual(product["discount_percent"], 0.25)
        old = {"woolworths:5": {**product, "promotional_price": None,
                                 "original_price": None, "discount_percent": None}}
        events = compare(old, {"woolworths:5": product}, "now")
        self.assertEqual(events[0]["change_type"], "Promotion")
        self.assertIn("2 for $6", render_html(events))

    def test_woolworths_availability_mapping(self):
        _, temporary = WoolworthsScraper._product({
            "Stockcode": 3, "Name": "Example Passata 700g",
            "IsAvailable": False, "IsInStock": False
        })
        self.assertEqual(temporary["availability_state"], "temporary_unavailable")
        _, out = WoolworthsScraper._product({
            "Stockcode": 4, "Name": "Example Passata 700g",
            "IsAvailable": True, "IsInStock": False
        })
        self.assertEqual(out["availability_state"], "out_of_stock")


class OnlineOnlyChangeTests(unittest.TestCase):
    def test_status_change_is_reported(self):
        old = {"coles:1": {"name": "A Pesto", "price": 2.0, "size": "100g",
                           "image_url": "a", "online_only": False}}
        new = {"coles:1": {"retailer": "Coles", "name": "A Pesto", "price": 2.0,
                           "size": "100g", "image_url": "a", "online_only": True,
                           "product_url": "u"}}
        events = compare(old, new, "2026-01-01T00:00:00+00:00")
        self.assertEqual(len(events), 1)
        self.assertEqual(events[0]["change_type"], "Online only")
        self.assertTrue(events[0]["online_only"])


class EmailVisibilityTests(unittest.TestCase):
    def test_promotion_ending_is_retained_but_hidden_from_email(self):
        old = {"coles:1": {"retailer": "Coles", "name": "Example Passata",
                            "price": 3.0, "original_price": 4.0,
                            "promotional_price": 3.0, "discount_percent": 0.25,
                            "size": "700g", "image_url": "", "product_url": "u"}}
        new = {"coles:1": {**old["coles:1"], "price": 4.0,
                            "original_price": None, "promotional_price": None,
                            "discount_percent": None}}
        events = compare(old, new, "now")
        self.assertEqual(len(events), 1)
        self.assertTrue(events[0]["promotion_ended"])
        self.assertEqual(email_visible_events(events), [])
        self.assertNotIn("Example Passata", render_html(events))


class AvailabilityLifecycleTests(unittest.TestCase):
    def test_temporary_unavailable_once_then_back_in_stock(self):
        temporary = {"1": {"retailer": "Woolworths", "name": "A Passata",
                            "availability_state": "temporary_unavailable",
                            "availability_label": "Temporarily unavailable",
                            "product_url": "u"}}
        first = compare({}, temporary, "now")
        self.assertEqual(first[0]["change_type"], "Unavailable")
        self.assertEqual(compare(temporary, temporary, "later"), [])
        self.assertEqual(visible_products({}, temporary), temporary)
        self.assertEqual(visible_products(temporary, temporary), {})
        available = {"1": {**temporary["1"], "availability_state": "in_stock",
                           "availability_label": "Available"}}
        back = compare(temporary, available, "later")
        self.assertEqual(back[0]["change_type"], "Restocked")
        self.assertEqual(visible_products(temporary, available), available)

    def test_out_of_stock_is_hidden(self):
        out = {"1": {"name": "A Passata", "availability_state": "out_of_stock"}}
        self.assertEqual(compare({}, out, "now"), [])
        self.assertEqual(visible_products({}, out), {})


class ChangeTests(unittest.TestCase):
    def test_changed_fields_new_products_and_deduplication(self):
        old = {"1": {"name": "A Pesto", "price": 2.0, "size": "100g", "image_url": "a"}}
        new = {
            "1": {"name": "A Pesto", "price": 2.5, "size": "100g", "image_url": "b", "product_url": "u"},
            "2": {"name": "B Passata", "price": 3.0, "size": "700g", "image_url": "c", "product_url": "v"},
        }
        events = compare(old, new, "2026-01-01T00:00:00+00:00")
        self.assertEqual([e["change_type"] for e in events],
                         ["RRP changed; Image 1 changed", "New"])
        self.assertEqual(len({e["product_id"] for e in events}), len(events))
        self.assertEqual(compare(old, new, "later", [e["event_id"] for e in events]), [])

    def test_price_summaries_distinguish_rrp_and_promotion(self):
        base = {"name": "A Pesto", "size": "100g", "image_url": "a",
                "product_url": "u"}
        rrp = compare({"1": {**base, "price": 4.0}},
                      {"1": {**base, "price": 5.0}}, "now")
        self.assertEqual(rrp[0]["change_type"], "RRP changed")
        promotion = compare(
            {"1": {**base, "price": 4.0, "original_price": None,
                    "promotional_price": None, "discount_percent": None}},
            {"1": {**base, "price": 3.0, "original_price": 4.0,
                    "promotional_price": 3.0, "discount_percent": 0.25}}, "later")
        self.assertEqual(promotion[0]["change_type"], "Promotion")

    def test_image_change_names_the_positions(self):
        base = {"name": "A Pesto", "price": 4.0, "size": "100g", "product_url": "u"}
        old = {"1": {**base, "image_url": "a", "image_urls": ["a", "b", "c"]}}
        new = {"1": {**base, "image_url": "a", "image_urls": ["a", "d", "c", "e"]}}
        events = compare(old, new, "now")
        self.assertEqual(events[0]["change_type"], "Image 2 changed; Image 4 added")

    def test_legacy_history_is_consolidated_per_sku_and_observation(self):
        events = [
            {"observed_at": "now", "product_id": "coles:1", "change_type": "Price changed",
             "event_id": "a", "name": "Pesto"},
            {"observed_at": "now", "product_id": "coles:1", "change_type": "Image changed",
             "event_id": "b", "name": "Pesto"},
        ]
        consolidated = consolidate_events(events)
        self.assertEqual(len(consolidated), 1)
        self.assertEqual(consolidated[0]["change_type"], "Price; Image")


if __name__ == "__main__":
    unittest.main()
